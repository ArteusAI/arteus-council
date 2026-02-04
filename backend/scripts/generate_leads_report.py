"""
Generate XLSX report of leads with their questions and answers.

This script connects to MongoDB, extracts all leads data, their conversations,
and generates a formatted XLSX report.

Usage:
    python generate_leads_report.py                     # All leads, all time
    python generate_leads_report.py 2024-01-01          # From date to now
    python generate_leads_report.py 2024-01-01 2024-12-31  # Date range
"""
import asyncio
from datetime import datetime
from typing import List, Dict, Any, Optional
import sys
import os
import argparse

# Add parent directory to path for imports
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from config import LEADS_MONGODB_URL, LEADS_MONGODB_DB_NAME

# Exclude this telegram user from reports
EXCLUDED_TELEGRAM = "@pavelfedortsov"


def get_mongo_client() -> AsyncIOMotorClient:
    """Create and return MongoDB client."""
    return AsyncIOMotorClient(LEADS_MONGODB_URL)


async def fetch_all_leads_data(
    db, 
    date_from: Optional[datetime] = None, 
    date_to: Optional[datetime] = None
) -> List[Dict[str, Any]]:
    """
    Fetch all leads and their conversations from MongoDB.
    
    Args:
        db: MongoDB database instance
        date_from: Filter conversations from this date (inclusive)
        date_to: Filter conversations until this date (inclusive)
    
    Returns:
        List of lead documents with embedded conversations
    """
    leads_collection = db["leads"]
    conversations_collection = db["conversations"]
    
    all_leads = []
    
    # Exclude specific telegram user
    lead_filter = {}
    if EXCLUDED_TELEGRAM:
        lead_filter["telegram"] = {"$ne": EXCLUDED_TELEGRAM}
    
    async for lead in leads_collection.find(lead_filter):
        telegram = lead.get("telegram", "")
        
        lead_data = {
            "session_id": lead.get("session_id"),
            "email": lead.get("email", ""),
            "telegram": telegram,
            "created_at": lead.get("created_at"),
            "template_id": lead.get("template_id", "default"),
            "conversations": []
        }
        
        # Build conversation filter
        conv_filter = {
            "session_id": lead_data["session_id"],
            "deleted_at": {"$exists": False}
        }
        
        # Add date filtering if specified
        if date_from or date_to:
            date_filter = {}
            if date_from:
                date_filter["$gte"] = date_from.isoformat()
            if date_to:
                # Include the entire day of date_to
                date_to_end = date_to.replace(hour=23, minute=59, second=59)
                date_filter["$lte"] = date_to_end.isoformat()
            
            if date_filter:
                conv_filter["created_at"] = date_filter
        
        # Fetch conversations for this lead
        async for conv in conversations_collection.find(conv_filter):
            lead_data["conversations"].append(conv)
        
        # Only include leads that have conversations in the date range
        if lead_data["conversations"]:
            all_leads.append(lead_data)
    
    return all_leads


def extract_qa_pairs(leads_data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Extract question-answer pairs from leads conversations.
    
    Args:
        leads_data: List of lead documents with conversations
    
    Returns:
        List of rows for XLSX report
    """
    rows = []
    
    for lead in leads_data:
        email = lead.get("email", "")
        telegram = lead.get("telegram", "")
        registration_date = lead.get("created_at")
        template_id = lead.get("template_id", "default")
        
        # Format registration date
        if isinstance(registration_date, datetime):
            reg_date_str = registration_date.strftime("%Y-%m-%d %H:%M:%S")
        else:
            reg_date_str = str(registration_date) if registration_date else ""
        
        for conversation in lead.get("conversations", []):
            conversation_id = conversation.get("_id", "")
            created_at = conversation.get("created_at", "")
            messages = conversation.get("messages", [])
            
            # Pair user questions with assistant answers
            for i, message in enumerate(messages):
                if message.get("role") == "user":
                    question = message.get("content", "")
                    
                    # Find corresponding assistant answer
                    answer = ""
                    if i + 1 < len(messages) and messages[i + 1].get("role") == "assistant":
                        stage3 = messages[i + 1].get("stage3", {})
                        if isinstance(stage3, dict):
                            answer = stage3.get("response", "")
                        else:
                            answer = str(stage3) if stage3 else ""
                    
                    # Add row only if we have a question
                    if question:
                        rows.append({
                            "email": email,
                            "telegram": telegram,
                            "registration_date": reg_date_str,
                            "template_id": template_id,
                            "question": question,
                            "answer": answer,
                            "conversation_id": str(conversation_id),
                            "asked_at": created_at
                        })
    
    return rows


def create_xlsx_report(rows: List[Dict[str, Any]], filename: str):
    """
    Create XLSX report from extracted data.
    
    Args:
        rows: List of data rows
        filename: Output filename
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Report"
    
    # Define headers
    headers = [
        "Lead Email",
        "Telegram",
        "Registration Date",
        "Template",
        "Question",
        "Final Answer",
        "Conversation ID",
        "Asked At"
    ]
    
    # Define column widths (in characters)
    column_widths = {
        "A": 25,  # Lead Email
        "B": 20,  # Telegram
        "C": 20,  # Registration Date
        "D": 15,  # Template
        "E": 50,  # Question
        "F": 50,  # Final Answer
        "G": 38,  # Conversation ID
        "H": 20,  # Asked At
    }
    
    # Set column widths
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    for row_num, row_data in enumerate(rows, 2):
        ws.cell(row=row_num, column=1, value=row_data.get("email", ""))
        ws.cell(row=row_num, column=2, value=row_data.get("telegram", ""))
        ws.cell(row=row_num, column=3, value=row_data.get("registration_date", ""))
        ws.cell(row=row_num, column=4, value=row_data.get("template_id", ""))
        ws.cell(row=row_num, column=5, value=row_data.get("question", ""))
        ws.cell(row=row_num, column=6, value=row_data.get("answer", ""))
        ws.cell(row=row_num, column=7, value=row_data.get("conversation_id", ""))
        ws.cell(row=row_num, column=8, value=row_data.get("asked_at", ""))
        
        # No text wrapping - single line cells
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
    
    # Save workbook
    wb.save(filename)
    print(f"✓ Report saved to: {filename}")


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Generate XLSX report of leads with their questions and answers.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python generate_leads_report.py                           # All leads, all time
  python generate_leads_report.py 2024-01-01                # From date to now
  python generate_leads_report.py "2024-01-01 10:30:00"     # From datetime to now
  python generate_leads_report.py 2024-01-01 2024-12-31     # Date range
  python generate_leads_report.py "2024-01-01 00:00:00" "2024-01-31 23:59:59"  # Datetime range

Date formats: YYYY-MM-DD or "YYYY-MM-DD HH:MM:SS"
        """
    )
    
    parser.add_argument(
        'date_from',
        nargs='?',
        help='Start date for filtering conversations (YYYY-MM-DD or "YYYY-MM-DD HH:MM:SS"). If not specified, includes all dates.'
    )
    
    parser.add_argument(
        'date_to',
        nargs='?',
        help='End date for filtering conversations (YYYY-MM-DD or "YYYY-MM-DD HH:MM:SS"). If not specified, includes up to now.'
    )
    
    return parser.parse_args()


def parse_date(date_str: str) -> Optional[datetime]:
    """
    Parse date string in multiple formats.
    
    Args:
        date_str: Date string in format YYYY-MM-DD or YYYY-MM-DD HH:MM:SS
    
    Returns:
        datetime object or None if parsing fails
    """
    if not date_str:
        return None
    
    # Try formats in order
    formats = [
        "%Y-%m-%d %H:%M:%S",  # Full datetime
        "%Y-%m-%d",           # Date only
    ]
    
    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    
    return None


async def main():
    """Main execution function."""
    # Parse command line arguments
    args = parse_arguments()
    
    # Parse dates
    date_from = None
    date_to = None
    
    if args.date_from:
        date_from = parse_date(args.date_from)
        if date_from is None:
            print(f"✗ Invalid date format for 'from' date: {args.date_from}")
            print('  Use format: YYYY-MM-DD or "YYYY-MM-DD HH:MM:SS"')
            return
    
    if args.date_to:
        date_to = parse_date(args.date_to)
        if date_to is None:
            print(f"✗ Invalid date format for 'to' date: {args.date_to}")
            print('  Use format: YYYY-MM-DD or "YYYY-MM-DD HH:MM:SS"')
            return
    
    # Validate date range
    if date_from and date_to and date_from > date_to:
        print("✗ 'From' date cannot be after 'to' date")
        return
    
    print("=" * 60)
    print("Leads Report Generator")
    print("=" * 60)
    
    # Show filter info
    if date_from or EXCLUDED_TELEGRAM:
        print("\nFilters:")
        if date_from:
            # Show full datetime if time component is present
            if date_from.hour != 0 or date_from.minute != 0 or date_from.second != 0:
                date_from_str = date_from.strftime("%Y-%m-%d %H:%M:%S")
            else:
                date_from_str = date_from.strftime("%Y-%m-%d")
            
            if date_to:
                if date_to.hour != 0 or date_to.minute != 0 or date_to.second != 0:
                    date_to_str = date_to.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    date_to_str = date_to.strftime("%Y-%m-%d")
                print(f"  Date range: {date_from_str} to {date_to_str}")
            else:
                print(f"  From date: {date_from_str} (to now)")
        if EXCLUDED_TELEGRAM:
            print(f"  Excluding telegram: {EXCLUDED_TELEGRAM}")
    
    # Connect to MongoDB
    print(f"\n1. Connecting to MongoDB...")
    print(f"   URL: {LEADS_MONGODB_URL}")
    print(f"   Database: {LEADS_MONGODB_DB_NAME}")
    
    client = get_mongo_client()
    db = client[LEADS_MONGODB_DB_NAME]
    
    try:
        # Test connection
        await client.admin.command('ping')
        print("   ✓ Connected successfully")
    except Exception as e:
        print(f"   ✗ Connection failed: {e}")
        return
    
    # Fetch leads data
    print("\n2. Fetching leads data...")
    try:
        leads_data = await fetch_all_leads_data(db, date_from, date_to)
        print(f"   ✓ Found {len(leads_data)} leads")
    except Exception as e:
        print(f"   ✗ Failed to fetch leads: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Extract Q&A pairs
    print("\n3. Extracting question-answer pairs...")
    try:
        rows = extract_qa_pairs(leads_data)
        print(f"   ✓ Extracted {len(rows)} Q&A pairs")
    except Exception as e:
        print(f"   ✗ Failed to extract data: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Generate filename with timestamp and date range
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    if date_from:
        date_suffix = f"_{date_from.strftime('%Y%m%d')}"
        if date_to:
            date_suffix += f"_{date_to.strftime('%Y%m%d')}"
        filename = f"leads_report{date_suffix}_{timestamp}.xlsx"
    else:
        filename = f"leads_report_{timestamp}.xlsx"
    
    # Create XLSX report
    print(f"\n4. Creating XLSX report...")
    try:
        create_xlsx_report(rows, filename)
    except Exception as e:
        print(f"   ✗ Failed to create report: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Summary
    print("\n" + "=" * 60)
    print("Summary:")
    print(f"  Total leads: {len(leads_data)}")
    print(f"  Total Q&A pairs: {len(rows)}")
    print(f"  Output file: {filename}")
    print("=" * 60)
    
    # Close MongoDB connection
    client.close()


if __name__ == "__main__":
    asyncio.run(main())
