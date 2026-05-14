"""
Standalone web dashboard for lead registrations.

Usage:
    uv run python backend/scripts/leads_dashboard.py
    uv run python backend/scripts/leads_dashboard.py --host 0.0.0.0 --port 8090

Environment:
    LEADS_MONGODB_URL       MongoDB URL, defaults to backend config value
    LEADS_MONGODB_DB_NAME   MongoDB database, defaults to backend config value
"""

from __future__ import annotations

import argparse
from io import BytesIO
import re
import secrets
import os
import sys
from datetime import datetime, timedelta, timezone
from typing import Any

from fastapi import Depends, FastAPI, HTTPException, Query
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from config import LEADS_MONGODB_DB_NAME, LEADS_MONGODB_URL  # noqa: E402


app = FastAPI(title="Leads Dashboard", docs_url=None, redoc_url=None)
client: AsyncIOMotorClient | None = None
security = HTTPBasic()


def _normalize_base_path(raw: str) -> str:
    """Return a base path starting with '/', without trailing slash and without duplicate slashes."""
    cleaned = (raw or "/leads").strip()
    if not cleaned.startswith("/"):
        cleaned = "/" + cleaned
    while "//" in cleaned:
        cleaned = cleaned.replace("//", "/")
    if len(cleaned) > 1 and cleaned.endswith("/"):
        cleaned = cleaned.rstrip("/")
    return cleaned


BASE_PATH = _normalize_base_path(os.getenv("LEADS_DASHBOARD_BASE_PATH", "/leads"))
LEADS_DASHBOARD_USERNAME = os.getenv("LEADS_DASHBOARD_USERNAME", "admin")
LEADS_DASHBOARD_PASSWORD = os.getenv("LEADS_DASHBOARD_PASSWORD", "arteus-leads")

URL_PATTERN = re.compile(
    r'(?:https?://)?(?:[a-zA-Z0-9-]+\.)+[a-zA-Z]{2,}(?:/[^\s<>\[\]()"\',;]*)?(?<![.,;:!?\)])',
    re.IGNORECASE,
)


def extract_first_url(text: str) -> str:
    """Return the first URL found in text with an https:// prefix if missing."""
    if not text:
        return ""
    match = URL_PATTERN.search(text)
    if not match:
        return ""
    url = match.group(0)
    if not url.startswith(("http://", "https://")):
        url = f"https://{url}"
    return url


def get_db():
    global client
    if client is None:
        client = AsyncIOMotorClient(os.getenv("LEADS_MONGODB_URL", LEADS_MONGODB_URL))
    db_name = os.getenv("LEADS_MONGODB_DB_NAME", LEADS_MONGODB_DB_NAME)
    return client[db_name]


def require_basic_auth(credentials: HTTPBasicCredentials) -> None:
    username_ok = secrets.compare_digest(credentials.username, LEADS_DASHBOARD_USERNAME)
    password_ok = secrets.compare_digest(credentials.password, LEADS_DASHBOARD_PASSWORD)
    if not username_ok or not password_ok:
        raise HTTPException(
            status_code=401,
            detail="Unauthorized",
            headers={"WWW-Authenticate": "Basic"},
        )


def to_iso(value: Any) -> str:
    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=timezone.utc)
        return value.isoformat()
    return str(value) if value else ""


def date_key(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str) and len(value) >= 10:
        return value[:10]
    return "unknown"


def contact_of(lead: dict[str, Any]) -> str:
    parts = [lead.get("telegram"), lead.get("email"), lead.get("linkedin")]
    contacts = [str(value).strip() for value in parts if value and str(value).strip()]
    return " / ".join(contacts) if contacts else "Без контакта"


def excluded_telegram_filter() -> dict[str, Any]:
    excluded_telegram = os.getenv("LEADS_DASHBOARD_EXCLUDED_TELEGRAM", "")
    if excluded_telegram:
        return {"telegram": {"$ne": excluded_telegram}}
    return {}


def serialize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return to_iso(value)
    if isinstance(value, list):
        return [serialize_value(item) for item in value]
    if isinstance(value, dict):
        return {str(key): serialize_value(item) for key, item in value.items()}
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


async def conversation_stats_by_session(session_ids: list[str]) -> dict[str, dict[str, Any]]:
    if not session_ids:
        return {}

    db = get_db()
    pipeline = [
        {
            "$match": {
                "session_id": {"$in": session_ids},
                "deleted_at": {"$exists": False},
            }
        },
        {"$sort": {"created_at": 1}},
        {
            "$group": {
                "_id": "$session_id",
                "conversation_count": {"$sum": 1},
                "message_count": {
                    "$sum": {
                        "$cond": [
                            {"$isArray": "$messages"},
                            {"$size": "$messages"},
                            0,
                        ]
                    }
                },
                "last_conversation_at": {"$max": "$created_at"},
                "first_messages": {"$first": "$messages"},
            }
        },
    ]

    stats: dict[str, dict[str, Any]] = {}
    async for row in db["conversations"].aggregate(pipeline):
        first_url = ""
        for message in row.get("first_messages") or []:
            if message.get("role") != "user":
                continue
            first_url = extract_first_url(message.get("content", "") or "")
            if first_url:
                break
        stats[row["_id"]] = {
            "conversation_count": row.get("conversation_count", 0),
            "message_count": row.get("message_count", 0),
            "last_conversation_at": to_iso(row.get("last_conversation_at")),
            "site_url": first_url,
        }
    return stats


async def conversations_for_session(session_id: str) -> list[dict[str, Any]]:
    db = get_db()
    conversations = []
    cursor = db["conversations"].find(
        {"session_id": session_id, "deleted_at": {"$exists": False}},
        {"_id": 1, "id": 1, "title": 1, "created_at": 1, "messages": 1},
    ).sort("created_at", 1)

    async for doc in cursor:
        conversation_id = doc.get("id") or doc.get("_id") or ""
        conversations.append(
            {
                "id": str(conversation_id),
                "title": doc.get("title", "New Conversation"),
                "created_at": to_iso(doc.get("created_at")),
                "messages": serialize_value(doc.get("messages", [])),
            }
        )

    return conversations


def report_rows_from_leads(leads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []

    for lead in leads:
        registration_date = lead.get("created_at")
        if isinstance(registration_date, datetime):
            registration_date = registration_date.strftime("%Y-%m-%d %H:%M:%S")
        else:
            registration_date = str(registration_date) if registration_date else ""

        site_url = ""
        for conversation in lead.get("conversations", []):
            for message in conversation.get("messages", []):
                if message.get("role") == "user":
                    site_url = extract_first_url(message.get("content", "") or "")
                    if site_url:
                        break
            if site_url:
                break

        for conversation in lead.get("conversations", []):
            conversation_id = conversation.get("_id") or conversation.get("id") or ""
            asked_at = to_iso(conversation.get("created_at"))
            messages = conversation.get("messages", [])

            for index, message in enumerate(messages):
                if message.get("role") != "user":
                    continue

                question = message.get("content", "")
                answer = ""
                if index + 1 < len(messages) and messages[index + 1].get("role") == "assistant":
                    stage3 = messages[index + 1].get("stage3", {})
                    if isinstance(stage3, dict):
                        answer = stage3.get("response", "")
                    elif stage3:
                        answer = str(stage3)

                if question:
                    rows.append(
                        {
                            "email": lead.get("email", ""),
                            "telegram": lead.get("telegram", ""),
                            "linkedin": lead.get("linkedin", ""),
                            "site_url": site_url,
                            "registration_date": registration_date,
                            "template_id": lead.get("template_id", "default"),
                            "question": question,
                            "answer": answer,
                            "conversation_id": str(conversation_id),
                            "asked_at": asked_at,
                        }
                    )

    return rows


def create_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Report"

    headers = [
        "Lead Email",
        "Telegram",
        "LinkedIn",
        "Site",
        "Registration Date",
        "Template",
        "Question",
        "Final Answer",
        "Conversation ID",
        "Asked At",
    ]
    column_widths = {
        "A": 25,
        "B": 20,
        "C": 30,
        "D": 35,
        "E": 20,
        "F": 15,
        "G": 50,
        "H": 50,
        "I": 38,
        "J": 20,
    }

    for column, width in column_widths.items():
        ws.column_dimensions[column].width = width

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_num, row_data in enumerate(rows, 2):
        values = [
            row_data.get("email", ""),
            row_data.get("telegram", ""),
            row_data.get("linkedin", ""),
            row_data.get("site_url", ""),
            row_data.get("registration_date", ""),
            row_data.get("template_id", ""),
            row_data.get("question", ""),
            row_data.get("answer", ""),
            row_data.get("conversation_id", ""),
            row_data.get("asked_at", ""),
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@app.get("/", include_in_schema=False)
async def root_redirect() -> RedirectResponse:
    return RedirectResponse(url=BASE_PATH)


@app.get(BASE_PATH, response_class=HTMLResponse)
@app.get(f"{BASE_PATH}/", response_class=HTMLResponse)
async def index(credentials: HTTPBasicCredentials = Depends(security)) -> str:
    require_basic_auth(credentials)
    return DASHBOARD_HTML.replace("__BASE_PATH__", BASE_PATH)


@app.get(f"{BASE_PATH}/api/leads")
async def leads_api(
    credentials: HTTPBasicCredentials = Depends(security),
    days: int = Query(default=30, ge=1, le=365),
    limit: int = Query(default=200, ge=1, le=1000),
):
    require_basic_auth(credentials)

    db = get_db()
    now = datetime.now(timezone.utc)
    since = now - timedelta(days=days)

    period_filter: dict[str, Any] = {"created_at": {"$gte": since}}
    all_filter = excluded_telegram_filter()
    period_filter.update(all_filter)

    leads_collection = db["leads"]
    total_count = await leads_collection.count_documents(all_filter)
    period_count = await leads_collection.count_documents(period_filter)
    today_count = await leads_collection.count_documents(
        {**all_filter, "created_at": {"$gte": now.replace(hour=0, minute=0, second=0, microsecond=0)}}
    )
    week_count = await leads_collection.count_documents(
        {**all_filter, "created_at": {"$gte": now - timedelta(days=7)}}
    )

    cursor = leads_collection.find(period_filter, {"_id": 0}).sort("created_at", -1).limit(limit)
    leads = [lead async for lead in cursor]
    session_ids = [lead.get("session_id") for lead in leads if lead.get("session_id")]
    conversation_stats = await conversation_stats_by_session(session_ids)

    buckets: dict[str, int] = {}
    enriched_leads = []
    active_leads = 0
    total_messages = 0

    for lead in leads:
        created_key = date_key(lead.get("created_at"))
        buckets[created_key] = buckets.get(created_key, 0) + 1

        stats = conversation_stats.get(lead.get("session_id"), {})
        conversation_count = stats.get("conversation_count", 0)
        message_count = stats.get("message_count", 0)
        active_leads += 1 if conversation_count else 0
        total_messages += message_count

        enriched_leads.append(
            {
                "session_id": lead.get("session_id", ""),
                "contact": contact_of(lead),
                "email": lead.get("email") or "",
                "telegram": lead.get("telegram") or "",
                "linkedin": lead.get("linkedin") or "",
                "template_id": lead.get("template_id", "default"),
                "created_at": to_iso(lead.get("created_at")),
                "conversation_count": conversation_count,
                "message_count": message_count,
                "last_conversation_at": stats.get("last_conversation_at", ""),
                "site_url": stats.get("site_url", ""),
            }
        )

    chart = [
        {"date": (since + timedelta(days=offset)).date().isoformat(), "count": 0}
        for offset in range(days + 1)
    ]
    for point in chart:
        point["count"] = buckets.get(point["date"], 0)

    return JSONResponse(
        {
            "generated_at": now.isoformat(),
            "database": os.getenv("LEADS_MONGODB_DB_NAME", LEADS_MONGODB_DB_NAME),
            "days": days,
            "total_count": total_count,
            "period_count": period_count,
            "today_count": today_count,
            "week_count": week_count,
            "active_leads": active_leads,
            "total_messages": total_messages,
            "chart": chart,
            "leads": enriched_leads,
        }
    )


@app.get(f"{BASE_PATH}/api/leads/{{session_id}}/conversations")
async def lead_conversations_api(
    session_id: str,
    credentials: HTTPBasicCredentials = Depends(security),
):
    require_basic_auth(credentials)

    db = get_db()
    lead_filter = {"session_id": session_id, **excluded_telegram_filter()}
    lead = await db["leads"].find_one(lead_filter, {"_id": 0})
    if lead is None:
        raise HTTPException(status_code=404, detail="Lead not found")

    return JSONResponse(
        {
            "session_id": session_id,
            "contact": contact_of(lead),
            "email": lead.get("email") or "",
            "telegram": lead.get("telegram") or "",
            "linkedin": lead.get("linkedin") or "",
            "template_id": lead.get("template_id", "default"),
            "created_at": to_iso(lead.get("created_at")),
            "conversations": await conversations_for_session(session_id),
        }
    )


@app.get(f"{BASE_PATH}/api/leads/export")
async def leads_export_api(
    credentials: HTTPBasicCredentials = Depends(security),
    days: int = Query(default=30, ge=1, le=365),
):
    require_basic_auth(credentials)

    db = get_db()
    now = datetime.now(timezone.utc)
    since = now - timedelta(days=days)
    lead_filter: dict[str, Any] = {"created_at": {"$gte": since}}
    lead_filter.update(excluded_telegram_filter())

    leads = []
    async for lead in db["leads"].find(lead_filter).sort("created_at", -1):
        lead["conversations"] = []
        cursor = db["conversations"].find(
            {"session_id": lead.get("session_id"), "deleted_at": {"$exists": False}}
        ).sort("created_at", 1)
        async for conversation in cursor:
            lead["conversations"].append(conversation)
        if lead["conversations"]:
            leads.append(lead)

    rows = report_rows_from_leads(leads)
    content = create_xlsx_bytes(rows)
    from_date = since.date().strftime("%Y%m%d")
    to_date = now.date().strftime("%Y%m%d")
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    filename = f"leads_report_{from_date}_{to_date}_{timestamp}.xlsx"

    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


DASHBOARD_HTML = """<!doctype html>
<html lang="ru">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Leads Dashboard</title>
  <style>
    :root {
      color-scheme: light;
      --bg: #f5f7fb;
      --panel: #ffffff;
      --text: #172033;
      --muted: #667085;
      --line: #d9e0ea;
      --accent: #136f63;
      --accent-soft: #dff4ef;
      --warn: #a15c00;
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font: 14px/1.45 Inter, ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
      color: var(--text);
      background: var(--bg);
    }
    header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 16px;
      padding: 22px 28px;
      border-bottom: 1px solid var(--line);
      background: var(--panel);
    }
    h1 { margin: 0; font-size: 22px; font-weight: 750; letter-spacing: 0; }
    .sub { color: var(--muted); margin-top: 3px; }
    .controls { display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
    select, button, input {
      height: 36px;
      border: 1px solid var(--line);
      background: #fff;
      color: var(--text);
      border-radius: 6px;
      padding: 0 10px;
      font: inherit;
    }
    button {
      cursor: pointer;
      border-color: var(--accent);
      background: var(--accent);
      color: #fff;
      font-weight: 650;
    }
    .ghost-btn {
      border-color: var(--line);
      background: #fff;
      color: var(--text);
      font-weight: 650;
    }
    .row-btn {
      height: 28px;
      padding: 0 8px;
      font-size: 12px;
      white-space: nowrap;
    }
    main { padding: 22px 28px 32px; }
    .grid {
      display: grid;
      grid-template-columns: repeat(5, minmax(130px, 1fr));
      gap: 12px;
      margin-bottom: 16px;
    }
    .metric, .chart-wrap, .table-wrap {
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
    }
    .metric { padding: 16px; min-height: 92px; }
    .label { color: var(--muted); font-size: 12px; text-transform: uppercase; letter-spacing: .04em; }
    .value { font-size: 30px; font-weight: 780; margin-top: 8px; }
    .chart-wrap { padding: 16px; margin-bottom: 16px; }
    .chart-head, .table-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      margin-bottom: 12px;
    }
    h2 { margin: 0; font-size: 16px; }
    .bars {
      display: grid;
      grid-auto-flow: column;
      grid-auto-columns: minmax(12px, 1fr);
      align-items: end;
      height: 190px;
      gap: 4px;
      padding-top: 10px;
      border-bottom: 1px solid var(--line);
    }
    .bar {
      min-height: 2px;
      background: var(--accent);
      border-radius: 4px 4px 0 0;
      position: relative;
    }
    .bar:hover::after {
      content: attr(data-label);
      position: absolute;
      left: 50%;
      bottom: calc(100% + 8px);
      transform: translateX(-50%);
      white-space: nowrap;
      background: #172033;
      color: #fff;
      padding: 5px 7px;
      border-radius: 5px;
      font-size: 12px;
      z-index: 2;
    }
    .table-wrap { overflow: hidden; }
    .table-head { padding: 16px 16px 0; }
    table { width: 100%; border-collapse: collapse; }
    th, td {
      text-align: left;
      padding: 11px 16px;
      border-top: 1px solid var(--line);
      vertical-align: top;
    }
    th { color: var(--muted); font-size: 12px; font-weight: 700; background: #fbfcfe; }
    .contact { font-weight: 680; }
    .mono { font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; font-size: 12px; color: var(--muted); }
    .site-cell { max-width: 260px; }
    .site-link {
      color: var(--accent);
      text-decoration: none;
      word-break: break-all;
      display: inline-block;
      max-width: 100%;
    }
    .site-link:hover { text-decoration: underline; }
    .empty, .error { padding: 20px; color: var(--muted); }
    .error { color: var(--warn); }
    .details-cell { background: #fbfcfe; padding: 0; }
    .details-panel { padding: 16px; border-top: 1px solid var(--line); }
    .conversation {
      background: #fff;
      border: 1px solid var(--line);
      border-radius: 8px;
      margin-bottom: 12px;
      overflow: hidden;
    }
    .conversation:last-child { margin-bottom: 0; }
    .conversation-head {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 12px;
      padding: 10px 12px;
      background: #f7fafc;
      border-bottom: 1px solid var(--line);
    }
    .conversation-title { font-weight: 700; }
    .message {
      display: grid;
      grid-template-columns: 90px minmax(0, 1fr);
      gap: 12px;
      padding: 12px;
      border-top: 1px solid #eef2f6;
    }
    .message:first-of-type { border-top: 0; }
    .role {
      color: var(--muted);
      font-weight: 700;
      text-transform: uppercase;
      font-size: 11px;
      letter-spacing: .04em;
    }
    .message-text {
      white-space: pre-wrap;
      overflow-wrap: anywhere;
    }
    @media (max-width: 900px) {
      header { align-items: flex-start; flex-direction: column; padding: 18px; }
      main { padding: 18px; }
      .grid { grid-template-columns: repeat(2, minmax(130px, 1fr)); }
      .table-wrap { overflow-x: auto; }
      table { min-width: 820px; }
    }
  </style>
</head>
<body>
  <header>
    <div>
      <h1>Заходы лидов</h1>
      <div class="sub" id="meta">Загрузка данных...</div>
    </div>
    <div class="controls">
      <select id="days">
        <option value="7">7 дней</option>
        <option value="30" selected>30 дней</option>
        <option value="90">90 дней</option>
        <option value="365">365 дней</option>
      </select>
      <button id="refresh" type="button">Обновить</button>
      <button id="export" class="ghost-btn" type="button">Выгрузить XLSX</button>
    </div>
  </header>
  <main>
    <section class="grid">
      <div class="metric"><div class="label">Всего</div><div class="value" id="total">-</div></div>
      <div class="metric"><div class="label">Сегодня</div><div class="value" id="today">-</div></div>
      <div class="metric"><div class="label">7 дней</div><div class="value" id="week">-</div></div>
      <div class="metric"><div class="label">За период</div><div class="value" id="period">-</div></div>
      <div class="metric"><div class="label">С диалогами</div><div class="value" id="active">-</div></div>
    </section>
    <section class="chart-wrap">
      <div class="chart-head">
        <h2>Динамика регистраций</h2>
        <div class="sub" id="messages">-</div>
      </div>
      <div class="bars" id="bars"></div>
    </section>
    <section class="table-wrap">
      <div class="table-head">
        <h2>Последние лиды</h2>
        <div class="sub" id="count">-</div>
      </div>
      <table>
        <thead>
          <tr>
            <th>Контакт</th>
            <th>Сайт</th>
            <th>Создан</th>
            <th>Диалоги</th>
            <th>Сообщения</th>
            <th>Последний диалог</th>
            <th>Session</th>
            <th></th>
          </tr>
        </thead>
        <tbody id="rows"></tbody>
      </table>
    </section>
  </main>
  <script>
    const $ = (id) => document.getElementById(id);
    const fmt = new Intl.DateTimeFormat('ru-RU', { dateStyle: 'short', timeStyle: 'short' });
    const expandedSessions = new Set();
    const conversationCache = new Map();
    let latestLeads = [];

    function formatDate(value) {
      if (!value) return '-';
      const date = new Date(value);
      return Number.isNaN(date.getTime()) ? value : fmt.format(date);
    }

    function renderBars(chart) {
      const max = Math.max(1, ...chart.map((point) => point.count));
      $('bars').innerHTML = chart.map((point) => {
        const height = Math.max(2, Math.round((point.count / max) * 180));
        return `<div class="bar" style="height:${height}px" data-label="${point.date}: ${point.count}"></div>`;
      }).join('');
    }

    function renderSite(url) {
      if (!url) return '<span class="sub">-</span>';
      const safe = escapeHtml(url);
      return `<a class="site-link" href="${safe}" target="_blank" rel="noopener noreferrer">${safe}</a>`;
    }

    function renderRows(leads) {
      latestLeads = leads;
      if (!leads.length) {
        $('rows').innerHTML = '<tr><td class="empty" colspan="8">За выбранный период лидов нет</td></tr>';
        return;
      }
      $('rows').innerHTML = leads.map((lead) => {
        const isExpanded = expandedSessions.has(lead.session_id);
        const buttonText = isExpanded ? 'Свернуть' : 'Диалог';
        const mainRow = `
        <tr>
          <td><div class="contact">${escapeHtml(lead.contact)}</div><div class="mono">${escapeHtml(lead.template_id)}</div></td>
          <td class="site-cell">${renderSite(lead.site_url)}</td>
          <td>${formatDate(lead.created_at)}</td>
          <td>${lead.conversation_count}</td>
          <td>${lead.message_count}</td>
          <td>${formatDate(lead.last_conversation_at)}</td>
          <td class="mono">${escapeHtml(lead.session_id)}</td>
          <td><button class="ghost-btn row-btn" type="button" data-session="${escapeHtml(lead.session_id)}">${buttonText}</button></td>
        </tr>
        `;
        if (!isExpanded) return mainRow;
        return `${mainRow}
          <tr>
            <td class="details-cell" colspan="8">
              <div class="details-panel" id="details-${escapeAttr(lead.session_id)}">${renderDetails(lead.session_id)}</div>
            </td>
          </tr>
        `;
      }).join('');
    }

    function escapeHtml(value) {
      return String(value ?? '').replace(/[&<>"']/g, (char) => ({
        '&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#39;'
      }[char]));
    }

    function escapeAttr(value) {
      return String(value ?? '').replace(/[^a-zA-Z0-9_-]/g, '_');
    }

    function assistantText(message) {
      if (message.stage3 && typeof message.stage3 === 'object' && message.stage3.response) {
        return message.stage3.response;
      }
      if (message.content) return message.content;
      return JSON.stringify(message, null, 2);
    }

    function renderMessage(message) {
      const role = message.role === 'user' ? 'Лид' : message.role === 'assistant' ? 'Ответ' : message.role || 'Сообщение';
      const text = message.role === 'assistant' ? assistantText(message) : message.content || JSON.stringify(message, null, 2);
      return `
        <div class="message">
          <div class="role">${escapeHtml(role)}</div>
          <div class="message-text">${escapeHtml(text)}</div>
        </div>
      `;
    }

    function renderConversation(conversation) {
      const messages = Array.isArray(conversation.messages) ? conversation.messages : [];
      return `
        <div class="conversation">
          <div class="conversation-head">
            <div>
              <div class="conversation-title">${escapeHtml(conversation.title || 'Диалог')}</div>
              <div class="mono">${escapeHtml(conversation.id)}</div>
            </div>
            <div class="sub">${formatDate(conversation.created_at)}</div>
          </div>
          ${messages.length ? messages.map(renderMessage).join('') : '<div class="empty">Сообщений нет</div>'}
        </div>
      `;
    }

    function renderDetails(sessionId) {
      const state = conversationCache.get(sessionId);
      if (!state || state.status === 'loading') return '<div class="sub">Загрузка диалога...</div>';
      if (state.status === 'error') return `<div class="error">${escapeHtml(state.error)}</div>`;
      const conversations = state.data.conversations || [];
      if (!conversations.length) return '<div class="empty">Диалогов нет</div>';
      return conversations.map(renderConversation).join('');
    }

    async function toggleLead(sessionId) {
      if (expandedSessions.has(sessionId)) {
        expandedSessions.delete(sessionId);
        renderRows(latestLeads);
        return;
      }

      expandedSessions.add(sessionId);
      if (!conversationCache.has(sessionId)) {
        conversationCache.set(sessionId, { status: 'loading' });
        fetchConversations(sessionId);
      }
      renderRows(latestLeads);
    }

    async function fetchConversations(sessionId) {
      try {
        const res = await fetch(`__BASE_PATH__/api/leads/${encodeURIComponent(sessionId)}/conversations`);
        if (!res.ok) throw new Error(`${res.status} ${res.statusText}`);
        const data = await res.json();
        conversationCache.set(sessionId, { status: 'loaded', data });
      } catch (error) {
        conversationCache.set(sessionId, { status: 'error', error: error.message });
      }
      renderRows(latestLeads);
    }

    function exportReport() {
      const days = $('days').value;
      window.location.href = `__BASE_PATH__/api/leads/export?days=${encodeURIComponent(days)}`;
    }

    async function load() {
      const days = $('days').value;
      $('meta').textContent = 'Загрузка данных...';
      try {
        const res = await fetch(`__BASE_PATH__/api/leads?days=${days}&limit=300`);
        if (!res.ok) throw new Error(`${res.status} ${res.statusText}`);
        const data = await res.json();
        $('total').textContent = data.total_count;
        $('today').textContent = data.today_count;
        $('week').textContent = data.week_count;
        $('period').textContent = data.period_count;
        $('active').textContent = data.active_leads;
        $('messages').textContent = `${data.total_messages} сообщений за период`;
        $('count').textContent = `${data.leads.length} строк`;
        $('meta').textContent = `${data.database} · обновлено ${formatDate(data.generated_at)}`;
        renderBars(data.chart);
        renderRows(data.leads);
      } catch (error) {
        $('meta').textContent = 'Ошибка загрузки';
        $('rows').innerHTML = `<tr><td class="error" colspan="8">${escapeHtml(error.message)}</td></tr>`;
      }
    }

    $('refresh').addEventListener('click', load);
    $('days').addEventListener('change', load);
    $('export').addEventListener('click', exportReport);
    $('rows').addEventListener('click', (event) => {
      const button = event.target.closest('button[data-session]');
      if (button) toggleLead(button.dataset.session);
    });
    load();
    setInterval(load, 5000);
  </script>
</body>
</html>
"""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run standalone leads dashboard.")
    parser.add_argument("--host", default=os.getenv("LEADS_DASHBOARD_HOST", "0.0.0.0"))
    parser.add_argument("--port", type=int, default=int(os.getenv("LEADS_DASHBOARD_PORT", "8090")))
    parser.add_argument("--reload", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    import uvicorn

    args = parse_args()
    target = "leads_dashboard:app" if args.reload else app
    uvicorn.run(target, host=args.host, port=args.port, reload=args.reload)
